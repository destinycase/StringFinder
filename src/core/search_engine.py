import os
import re
import mmap
import bisect
from os.path import splitext
import logging
import unicodedata

# 로거 설정
logger = logging.getLogger("StringFinder.SearchEngine")


def normalize_unicode(text):
    """문자열을 NFC 방식으로 정규화합니다."""
    if text is None:
        return ""
    return unicodedata.normalize("NFC", str(text))


def detect_encoding_quickly(data):
    """
    바이트 데이터의 인코딩을 빠르게 판별합니다.
    주로 UTF-8과 CP949(EUC-KR)를 구분하는 데 사용됩니다.
    """
    if not data:
        return "utf-8"

    # BOM 확인
    if data.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    if data.startswith(b"\xff\xfe"):
        return "utf-16"
    if data.startswith(b"\xfe\xff"):
        return "utf-16-be"

    # UTF-8 시도
    try:
        data.decode("utf-8")
        return "utf-8"
    except UnicodeDecodeError:
        pass

    # CP949 시도
    try:
        data.decode("cp949")
        return "cp949"
    except UnicodeDecodeError:
        pass

    return "utf-8"  # 기본값


class SearchEngine:
    """
    검색 환경 및 전반적인 설정을 관리하는 마스터 엔진 클래스입니다.
    현재는 주요 검색 로직이 최적화를 위해 전역 함수로 분리되어 있으나,
    향후 확장성을 위해 인스턴스 형태로 유지합니다.
    """

    def __init__(self):
        pass


class FileScanner:
    """
    검색 대상이 되는 파일 리스트를 효율적으로 스캔하고 필터링하는 클래스입니다.
    """

    def __init__(self, folders, extensions, filename_filter=None, stop_check_callback=None):
        """
        스캐너를 초기화합니다.

        Args:
            folders (list): 검색할 폴더 경로 리스트
            extensions (list): 검색할 확장자 리스트 (예: ['.txt', '.py'])
            filename_filter (str): 파일명에 포함되어야 할 필터링 문자열
            stop_check_callback (callable): 중단 여부를 반환하는 함수 (True 반환 시 중단)
        """
        self.folders = folders
        self.extensions = [(e if e.startswith(".") else "." + e).lower() for e in extensions]
        self.filename_filter = filename_filter.lower() if filename_filter else None
        self.stop_check_callback = stop_check_callback

    def scan(self):
        """
        설정된 폴더들을 순회하며 조건에 맞는 파일 리스트를 수집합니다.
        os.scandir를 사용하여 파일 속성(크기) 접근 시 시스템 콜을 최소화합니다.
        """
        file_list = []
        for folder in self.folders:
            if self.stop_check_callback and self.stop_check_callback():
                break

            if not os.path.exists(folder):
                continue

            self._scan_recursive(folder, file_list)
        return file_list

    def _scan_recursive(self, folder, file_list):
        """
        재귀적으로 폴더를 스캔하며 파일 정보를 수집합니다.
        """
        try:
            # os.scandir는 컨텍스트 매니저로 사용하여 리소스를 확실히 해제합니다.
            with os.scandir(folder) as entries:
                for entry in entries:
                    if self.stop_check_callback and self.stop_check_callback():
                        return

                    try:
                        # 심볼릭 링크는 보안 및 무한 루프 방지를 위해 기본적으로 따라가지 않거나 상황에 맞게 처리
                        # 여기서는 is_dir(), is_file() 호출 시 follow_symlinks=False를 명시하지 않으면 기본값(True)일 수 있음
                        # 성능을 위해 stat() 호출 최소화
                        if entry.is_dir():
                            self._scan_recursive(entry.path, file_list)
                        elif entry.is_file():
                            ext = splitext(entry.name)[1].lower()
                            if ext in self.extensions:
                                if self.filename_filter and self.filename_filter not in entry.name.lower():
                                    continue

                                # os.scandir의 entry.stat()은 윈도우에서 추가 시스템 콜 없이 캐시된 정보를 반환함
                                size = entry.stat().st_size
                                file_list.append((entry.path, size))
                    except (OSError, PermissionError):
                        continue
        except (OSError, PermissionError):
            # 폴더 접근 권한이 없는 경우 무시
            pass


def strip_comments(content, ext):
    """
    파일 확장자에 따라 주석을 제거하고 공백으로 대체합니다 (줄 번호 유지).
    """

    def replacer(match):
        # 개행 문자를 제외한 모든 문자를 공백으로 치환하여 위치와 줄 번호를 보존함
        return re.sub(r"[^\n\r]", " ", match.group())

    if ext in [".json", ".js", ".c", ".cpp", ".cs", ".java"]:
        # // 또는 /* */ 주석 제거
        return re.sub(r"//.*?\n|/\*.*?\*/", replacer, content, flags=re.S)
    elif ext in [".xml", ".html"]:
        # <!-- --> 주석 제거
        return re.sub(r"<!--.*?-->", replacer, content, flags=re.S)
    elif ext in [".py", ".rb", ".sh", ".yaml", ".yml"]:
        # # 주석 제거
        return re.sub(r"#.*?\n", replacer, content)
    elif ext in [".sql"]:
        # -- 주석 제거
        return re.sub(r"--.*?\n", replacer, content)
    return content


def search_in_excel(file_path, search_string):
    """
    Excel 파일(.xlsx, .xlsm 등) 내의 모든 시트에서 문자열을 검색합니다.
    (엑셀은 주석 개념이 셀 내용과 분리되어 있어 기본적으로 값만 검색함)
    """
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, data_only=True, read_only=True)
        count = 0
        matches = []
        search_string = normalize_unicode(search_string).lower()

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        val = normalize_unicode(str(cell.value)).lower()
                        if search_string in val:
                            count += 1
                            pos = f"{sheet_name}!{cell.coordinate}"
                            matches.append((pos, str(cell.value)))
        wb.close()
        if count > 0:
            return (file_path, count, matches)
    except Exception as e:
        logger.debug(f"Excel search error ({file_path}): {e}")
        return "SKIPPED"
    return None


def search_in_json_special(file_path, search_string, exact_match=False):
    """
    JSON 파일을 파싱하여 값(Value)들 중에서만 검색합니다. 주석은 무시합니다.
    """
    try:
        import json

        with open(file_path, "rb") as f:
            head = f.read(1024)
            encoding = detect_encoding_quickly(head)

        with open(file_path, "r", encoding=encoding, errors="replace") as f:
            raw_content = f.read()

        # 주석 제거 (공백 대체로 오프셋 보존)
        processed_content = strip_comments(raw_content, ".json")
        # 후행 쉼표 처리
        processed_content = re.sub(r",\s*([}\]])", lambda m: " " * (len(m.group()) - 1) + m.group(1), processed_content)

        try:
            data = json.loads(processed_content, strict=False)
        except json.JSONDecodeError:
            return "INVALID"

        matches = []
        search_string = normalize_unicode(search_string).lower()
        newline_indices = [i for i, char in enumerate(raw_content) if char == "\n"]

        def get_line_no(pos):
            import bisect

            return bisect.bisect_left(newline_indices, pos) + 1

        search_state = {"last_pos": 0}

        def _recursive_search(obj, path=""):
            count = 0
            if isinstance(obj, dict):
                for k, v in obj.items():
                    count += _recursive_search(v, str(k))
            elif isinstance(obj, list):
                for i, v in enumerate(obj):
                    count += _recursive_search(v, f"[{i}]")
            else:
                val_raw = normalize_unicode(str(obj))
                val_lower = val_raw.lower()

                is_match = False
                if exact_match:
                    is_match = val_lower == search_string
                else:
                    is_match = search_string in val_lower

                if is_match:
                    count += 1
                    # 위치 탐색 시 오탐 방지: 단어 경계(\b)를 활용하여 정확한 값의 위치를 찾음
                    # 값의 앞뒤가 영문/숫자가 아닌 경우에도 대응하기 위해 \b 사용
                    pattern_str = re.escape(val_raw)
                    if val_raw.isalnum() or (val_raw.startswith('"') and val_raw.endswith('"')):
                        pattern_str = rf"\b{pattern_str}\b"

                    pattern = re.compile(pattern_str, re.IGNORECASE)
                    match = pattern.search(processed_content, search_state["last_pos"])

                    if match:
                        line_no = get_line_no(match.start())
                        search_state["last_pos"] = match.end()
                    else:
                        match_retry = pattern.search(processed_content)
                        if match_retry:
                            line_no = get_line_no(match_retry.start())
                            search_state["last_pos"] = match_retry.end()
                        else:
                            line_no = 1

                    matches.append((line_no, path or "root", val_raw))
            return count

        total_count = _recursive_search(data)
        if total_count > 0:
            return (file_path, total_count, matches)
        return []  # 발견된 것 없음
    except Exception as e:
        logger.debug(f"JSON special search error ({file_path}): {e}")
        return "INVALID"


def search_in_xml_special(file_path, search_string, exact_match=False):
    """
    XML 파일을 파싱하여 검색합니다. 주석은 무시합니다.
    """
    try:
        import xml.parsers.expat

        # 주석 제거를 위해 먼저 읽음
        with open(file_path, "rb") as f:
            head = f.read(1024)
            encoding = detect_encoding_quickly(head)

        with open(file_path, "r", encoding=encoding, errors="replace") as f:
            raw_content = f.read()
        processed_content = strip_comments(raw_content, ".xml")

        matches = []
        count = 0
        search_string = normalize_unicode(search_string).lower()

        class XMLSearcher:
            def __init__(self):
                self.parser = xml.parsers.expat.ParserCreate()
                self.parser.StartElementHandler = self.start_element
                self.parser.CharacterDataHandler = self.char_data
                self.current_tags = []

            def start_element(self, name, attrs):
                nonlocal count
                self.current_tags.append(name)
                for k, v in attrs.items():
                    val_str = normalize_unicode(str(v))
                    val_lower = val_str.lower()

                    is_match = (val_lower == search_string) if exact_match else (search_string in val_lower)
                    if is_match:
                        count += 1
                        line = self.parser.CurrentLineNumber
                        matches.append((line, str(k), val_str))

            def char_data(self, data):
                nonlocal count
                text = normalize_unicode(data).strip()
                if text:
                    text_lower = text.lower()
                    is_match = (text_lower == search_string) if exact_match else (search_string in text_lower)
                    if is_match:
                        count += 1
                        line = self.parser.CurrentLineNumber
                        name = self.current_tags[-1] if self.current_tags else "root"
                        matches.append((line, name, text))

            def parse_content(self, content):
                self.parser.Parse(content, True)

        searcher = XMLSearcher()
        try:
            searcher.parse_content(processed_content)
        except Exception:
            return "INVALID"

        if count > 0:
            return (file_path, count, matches)
        return []
    except Exception as e:
        logger.debug(f"XML special search error ({file_path}): {e}")
        return "INVALID"


def _quick_search_bytes(file_path, search_text):
    """
    파일 전체를 디코딩하지 않고 바이트 패턴 매칭으로 검색어 존재 여부를 빠르게 확인합니다.
    (UTF-8, CP949, UTF-16LE 인코딩을 고려)
    True를 반환하면 파일 내에 검색어(의 바이트 표현)가 존재할 가능성이 높습니다.
    """
    patterns = []
    # 일반적인 인코딩들에 대해 검색어 바이트열 생성
    # 순서: UTF-8 (가장 흔함) -> CP949 (한글 윈도우) -> UTF-16LE (윈도우 시스템 파일)
    encodings = ["utf-8", "cp949", "utf-16-le"]

    for enc in encodings:
        try:
            pat = search_text.encode(enc)
            if pat:
                patterns.append(pat)
        except UnicodeEncodeError:
            pass

    if not patterns:
        return False

    try:
        # 파일 크기 확인 (여기서는 os.path.getsize를 쓰지만, 호출처에서 넘겨준다면 더 좋음)
        # 0바이트 파일은 패스
        if os.path.getsize(file_path) == 0:
            return False

        with open(file_path, "rb") as f:
            # mmap을 사용하여 메모리 복사 없이 고속 검색
            with mmap.mmap(f.fileno(), 0, access=mmap.ACCESS_READ) as mm:
                for pat in patterns:
                    if mm.find(pat) >= 0:
                        return True
    except (ValueError, OSError) as e:
        # 파일에 접근조차 못하는 경우는 상위에서 SKIPPED 처리하도록 예외 다시 던짐
        raise e

    return False


def search_in_file(file_path, search_string, file_size=None, special_mode=None):
    """
    고성능 검색 엔진: mmap과 바이트 레벨 정규표현식을 조합하여 파일을 검색합니다.
    - 대량의 파일 검색 시 메모리 사용량을 최소화합니다.
    - 주석 제외 검색 및 XML/JSON 특수 모드를 지원합니다.
    - 파일의 인코딩을 자동으로 판별하여 처리합니다.
    """
    search_string_nfc = normalize_unicode(search_string)
    ext = splitext(file_path)[1].lower()

    # Excel 검색 (Binary Pre-check 제외: Zip 압축 파일이므로 바이트 매칭 불가)
    if ext in [".xlsx", ".xlsm", ".xls", ".xlsb"]:
        return search_in_excel(file_path, search_string_nfc)

    # [Optimization] Binary Pre-check
    # 파일을 디코딩하거나 파싱하기 전에, 바이트 단위로 검색어가 있는지 먼저 확인합니다.
    try:
        found_in_binary = _quick_search_bytes(file_path, search_string_nfc)
        if not found_in_binary:
            return None
    except Exception as e:
        logger.debug(f"Binary pre-check error ({file_path}): {e}")
        return "SKIPPED"

    # 특수 검색 모드 (XML/JSON)
    if special_mode:
        is_xml = "XML" in special_mode
        is_json = "JSON" in special_mode
        is_exact = "전체 일치" in special_mode

        res = None
        if is_xml and ext == ".xml":
            res = search_in_xml_special(file_path, search_string_nfc, exact_match=is_exact)
        elif is_json and ext == ".json":
            res = search_in_json_special(file_path, search_string_nfc, exact_match=is_exact)

        if res == "INVALID":
            return "SKIPPED"  # 규격 미준수 파일 스킵 신호
        if res and res != []:
            return res
        if res == []:
            return None  # 매치 없음

    # 일반 텍스트 검색 (주석 제외 요구사항 반영을 위해 전체 읽기 방식 병행)
    try:
        if file_size is None:
            file_size = os.path.getsize(file_path)

        if file_size == 0:
            return None

        # 대용량 파일이 아닐 경우(주석 제거를 위해 전체 읽기)
        if file_size < 10 * 1024 * 1024:
            with open(file_path, "rb") as f:
                head = f.read(1024)
                encoding = detect_encoding_quickly(head)

            with open(file_path, "r", encoding=encoding, errors="replace") as f:
                content = f.read()
            processed_content = strip_comments(content, ext)

            matches = []
            count = 0
            lines = processed_content.splitlines()
            search_low = search_string_nfc.lower()

            for i, line in enumerate(lines):
                if search_low in line.lower():
                    count += 1
                    matches.append((i + 1, line.strip()))

            if count > 0:
                return (file_path, count, matches)
            return None

        # 10MB 이상 대용량 파일은 mmap 유지 (주석 처리가 까다로움 - 일단 mmap으로 수행하되 주석이 많은 언어는 경고 처리 가능)
        with open(file_path, "rb") as f:
            with mmap.mmap(f.fileno(), 0, access=mmap.ACCESS_READ) as mm:
                search_pattern_bytes = search_string_nfc.encode("utf-8", errors="ignore")
                if not search_pattern_bytes:
                    return None

                flags = re.IGNORECASE
                pattern = re.escape(search_pattern_bytes)
                newline_offsets = [m.start() for m in re.finditer(b"\n", mm)]
                mm.seek(0)

                matches = []
                count = 0
                for match in re.finditer(pattern, mm, flags):
                    count += 1
                    start_pos = match.start()
                    line_no = bisect.bisect_right(newline_offsets, start_pos) + 1
                    line_start = 0 if line_no == 1 else newline_offsets[line_no - 2] + 1
                    line_end = newline_offsets[line_no - 1] if line_no <= len(newline_offsets) else file_size
                    # mmap은 바이트이므로 디코딩 필요. 큰 파일도 인코딩 감지를 해야 하지만,
                    # 성능상 UTF-8 우선 시도 후 실패시 replace 처리
                    line_content = mm[line_start:line_end].decode("utf-8", errors="replace").strip()
                    matches.append((line_no, line_content))

                if count > 0:
                    return (file_path, count, matches)
    except Exception as e:
        logger.debug(f"Search error in {file_path}: {e}")
        return "SKIPPED"
    return None


def search_in_files_batch(file_batch, search_string, special_mode=None):
    """
    배치 단위 검색을 수행하고 결과와 스킵된 목록을 반환합니다.
    """
    results = []
    skipped = []
    for f_path, f_size in file_batch:
        res = search_in_file(f_path, search_string, f_size, special_mode)
        if res == "SKIPPED":
            skipped.append(f_path)
        elif res:
            results.append(res)
    return {"results": results, "skipped": skipped}
