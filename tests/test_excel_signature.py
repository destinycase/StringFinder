"""
[test_excel_signature.py]

이 테스트는 Excel 파일의 유효성을 확장자가 아닌 바이너리 시그너처(Magic Number)를 통해 판별하는 로직을 검증합니다.

- 테스트 목적:
  1. 확장자만 엑셀로 변경된 일반 텍스트 파일이나 손상된 파일에 대한 검색 시도를 사전에 차단.
  2. 엑셀 특수 검색 엔진의 안정적인 진입점 확보 및 예외 방지.

- 주요 검증 사항:
  1. 표준 Excel(Zip 기반) 시그너처 (`\x50\x4b\x03\x04`) 감지 확인.
  2. 잘못된 시그너처를 가진 파일에 대한 'SKIPPED' 상태 반환 및 에러 메시지 무결성.
"""

import os
from typing import Any, List
from unittest.mock import patch

import pytest

from core.search_engine import (
    format_excel_panic_reason,
    is_valid_excel_signature,
    search_in_excel_special,
    search_in_file,
)
from sf_utils.app_strings import AppStrings
from sf_utils.constants import Constants


def test_excel_signature_check(tmp_path):
    with patch("core.search_engine.HAS_RUST_ENGINE", False):
        valid_xlsx = tmp_path / "valid.xlsx"
        with open(valid_xlsx, "wb") as f:
            f.write(b"PK\x03\x04" + b"\x00" * 10)
        assert is_valid_excel_signature(str(valid_xlsx))

        valid_xls = tmp_path / "valid.xls"
        with open(valid_xls, "wb") as f:
            f.write(b"\xd0\xcf\x11\xe0" + b"\x00" * 10)
        assert is_valid_excel_signature(str(valid_xls))

        invalid_file = tmp_path / "fake.xlsx"
        with open(invalid_file, "w") as f:
            f.write("This is just a text file")
        assert not is_valid_excel_signature(str(invalid_file))

        res = search_in_excel_special(str(invalid_file), "test", exact_match=False)
        assert res is not None
        res_list: List[Any] = list(res)
        assert res_list[0] == "SKIPPED"
        assert len(res_list) >= 2
        assert isinstance(res_list[1], str)


def test_excel_access_error_classification(tmp_path, monkeypatch):
    with patch("core.search_engine.HAS_RUST_ENGINE", False):
        target = tmp_path / "locked.xlsx"
        with open(target, "wb") as f:
            f.write(b"PK\x03\x04" + b"\x00" * 10)

        real_open = open

        def fake_open(path, mode="r", *args, **kwargs):
            if os.fspath(path) == str(target) and "rb" in mode:
                raise PermissionError("access denied")
            return real_open(path, mode, *args, **kwargs)

        monkeypatch.setattr("builtins.open", fake_open)
        res_special = search_in_excel_special(str(target), "keyword", exact_match=False)
        res_file = search_in_file(str(target), "keyword")

        assert isinstance(res_special, tuple) and res_special[0] == Constants.STATUS_SKIPPED
        assert isinstance(res_file, tuple) and res_file[0] == Constants.STATUS_SKIPPED
        access_prefix = AppStrings.ERROR_EXCEL_ACCESS.split(":")[0]
        assert access_prefix in str(res_special[1])


def test_rust_excel_panic_marker_returns_skipped(tmp_path):
    class _FakeMatch:
        def __init__(self, content):
            self.content = content
            self.line = 1

        def __getitem__(self, item):
            if item == 0:
                return self.line
            if item == 1:
                return self.content
            raise IndexError

    target = tmp_path / "panic.xlsx"
    with open(target, "wb") as f:
        f.write(b"PK\x03\x04" + b"\x00" * 10)

    with (
        patch("core.search_engine.HAS_RUST_ENGINE", True),
        patch(
            "core.search_engine.sf_engine.search_file",
            return_value=[_FakeMatch("__SF_EXCEL_PANIC__|xlsx")],
        ),
    ):
        res = search_in_excel_special(str(target), "keyword", exact_match=False)

    assert res is not None
    res_list: List[Any] = list(res)
    assert res_list[0] == Constants.STATUS_SKIPPED
    assert res_list[1] == format_excel_panic_reason("xlsx")


def test_rust_excel_sheet_error_marker_is_filtered(tmp_path):
    class _FakeMatch:
        def __init__(self, content):
            self.content = content
            self.line = 1

        def __getitem__(self, item):
            if item == 0:
                return self.line
            if item == 1:
                return self.content
            raise IndexError

    target = tmp_path / "sheet_error.xlsx"
    with open(target, "wb") as f:
        f.write(b"PK\x03\x04" + b"\x00" * 10)

    with (
        patch("core.search_engine.HAS_RUST_ENGINE", True),
        patch(
            "core.search_engine.sf_engine.search_file",
            return_value=[
                _FakeMatch("__SF_EXCEL_SHEET_ERR__|Sheet1|parse fail"),
                _FakeMatch("Sheet2 | A1 | value"),
            ],
        ),
    ):
        res = search_in_excel_special(str(target), "keyword", exact_match=False)

    assert res is not None
    res_list: List[Any] = list(res)
    assert int(res_list[1]) == 1
    matches: List[Any] = list(res_list[2])
    assert str(matches[0][1]) == "Sheet2"
    assert str(matches[0][2]) == "A1"


def test_rust_excel_existence_cell_limit_returns_skipped(tmp_path):
    target = tmp_path / "limited.xlsx"
    with open(target, "wb") as f:
        f.write(b"PK\x03\x04" + b"\x00" * 10)

    with (
        patch("core.search_engine.HAS_RUST_ENGINE", True),
        patch(
            "core.search_engine.sf_engine.search_file",
            return_value=[(0, "__SF_EXCEL_CELL_LIMIT__|2", None, None)],
        ),
    ):
        result = search_in_excel_special(
            str(target),
            "keyword",
            existence_only=True,
        )

    assert result == (
        Constants.STATUS_SKIPPED,
        AppStrings.SKIP_REASON_EXCEL_CELL_LIMIT.format(2),
    )


def test_python_excel_existence_cell_limit_returns_skipped(tmp_path, monkeypatch):
    from openpyxl import Workbook

    target = tmp_path / "limited_python.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["first", "second", "needle"])
    workbook.save(target)
    workbook.close()

    monkeypatch.setattr(
        "core.search_engine.ConfigManager.get_advanced_settings",
        lambda _self: {Constants.CONFIG_KEY_MAX_CHECK_CELLS: 2},
    )
    result = search_in_excel_special(
        str(target),
        "needle",
        use_complex_search=True,
        existence_only=True,
    )

    assert result == (
        Constants.STATUS_SKIPPED,
        AppStrings.SKIP_REASON_EXCEL_CELL_LIMIT.format(2),
    )


def test_python_excel_precise_search_keeps_single_cell_sheet(tmp_path):
    from openpyxl import Workbook

    target = tmp_path / "single_cell.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "needle"
    workbook.save(target)
    workbook.close()

    result = search_in_excel_special(
        str(target),
        "needle",
        use_complex_search=True,
    )

    assert isinstance(result, tuple) and len(result) == 3
    assert result[1] == 1
    assert result[2][0][2] == "A1"


def test_real_rust_excel_existence_limit_and_early_match(tmp_path, monkeypatch):
    from core import search_engine
    from openpyxl import Workbook

    if not search_engine.HAS_RUST_ENGINE:
        pytest.skip("compiled Rust engine is unavailable")

    limited_target = tmp_path / "rust_limited.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["first", "second", "needle"])
    workbook.save(limited_target)
    workbook.close()

    matched_target = tmp_path / "rust_early_match.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["needle", "second", "third"])
    workbook.save(matched_target)
    workbook.close()

    monkeypatch.setattr(
        search_engine.ConfigManager,
        "get_advanced_settings",
        lambda _self: {Constants.CONFIG_KEY_MAX_CHECK_CELLS: 2},
    )

    limited = search_in_excel_special(
        str(limited_target),
        "needle",
        existence_only=True,
    )
    matched = search_in_excel_special(
        str(matched_target),
        "needle",
        existence_only=True,
    )

    assert limited == (
        Constants.STATUS_SKIPPED,
        AppStrings.SKIP_REASON_EXCEL_CELL_LIMIT.format(2),
    )
    assert isinstance(matched, tuple) and len(matched) == 3
    assert matched[1] == 1
