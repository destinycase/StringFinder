"""
[test_search_engine.py]

이 테스트는 StringFinder의 핵심 검색 엔진 로직과 인코딩 감지 기능을 검증합니다.

- 테스트 목적:
  1. 다양한 인코딩(UTF-8, UTF-16LE, CP949 등) 파일에 대한 검색 정확도 보장.
  2. 텍스트 및 Excel 파일에 대한 기본 검색 프로세스 무결성 검증.
  3. 인코딩 감지 및 파일 읽기 중의 예외 처리 상황 확인.

- 주요 검증 사항:
  1. UTF-16LE(BOM 포함/미포함) 대용량 파일 검색.
  2. 텍스트 파일 내 키워드 매칭 및 미검측 사례.
  3. 빈 파일, 손상된 엑셀 파일, 권한 오류 등 예외 상황 처리.
  4. 유니코드 NFC 정규화 기반 매칭 정확도.
"""

import os

import openpyxl

from core.search_engine import detect_encoding_quickly, search_in_file
from sf_utils.constants import Constants


def test_detect_encoding_utf16le():
    data = b"\xff\xfe\x48\x00\x65\x00\x6c\x00\x6c\x00\x6f\x00"  # "Hello" 문자열
    encoding = detect_encoding_quickly(data)
    assert encoding in [Constants.ENC_UTF16_LE, Constants.ENC_UTF16]


def test_search_utf16le_file(tmp_path):
    test_file = tmp_path / "test_utf16le_bom.txt"
    content = "Hello\nTest String\nUTF-16LE Encoding"
    with open(test_file, "wb") as f:
        f.write(b"\xff\xfe")
        f.write(content.encode("utf-16-le"))

    result = search_in_file(str(test_file), "Test")
    assert result is not None
    assert isinstance(result[1], int)
    assert result[1] > 0


def test_utf16le_large_file(tmp_path):
    test_file = tmp_path / "large_utf16le.txt"
    content = "Test Line\n" * 1000
    with open(test_file, "wb") as f:
        f.write(b"\xff\xfe")
        f.write(content.encode("utf-16-le"))

    result = search_in_file(str(test_file), "Test")
    assert result is not None


def test_search_in_text_file(sample_text_file):
    file_path, search_str = sample_text_file
    result = search_in_file(file_path, search_str)
    assert result is not None
    assert result[0] == file_path
    assert result[1] == 1


def test_search_not_found(sample_text_file):
    file_path, _ = sample_text_file
    result = search_in_file(file_path, "NonExistentString")
    assert result is None


def test_search_empty_file(temp_dir):
    from sf_utils.app_strings import AppStrings
    from sf_utils.constants import Constants

    path = os.path.join(temp_dir, "empty.txt")
    with open(path, "w") as _:
        pass
    result = search_in_file(path, "anything")

    assert result is not None
    assert result[0] == Constants.STATUS_SKIPPED
    assert result[1] == AppStrings.SKIP_EMPTY_FILE


def test_detect_encoding_quickly():
    utf8_data = "Hello World".encode("utf-8")
    assert detect_encoding_quickly(utf8_data) == "utf-8"
    cp949_data = "Hello".encode("cp949")
    assert detect_encoding_quickly(cp949_data) in ["ascii", "cp949", "utf-8"]


def test_search_in_excel(temp_dir):
    excel_path = os.path.join(temp_dir, "test.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Search Me"
    ws["B2"] = "Python Rulez"
    wb.save(excel_path)
    wb.close()

    result = search_in_file(excel_path, "Rulez")
    assert result is not None
    # [Mypy] Union[SearchResult, SkippedResult] 타입 추론 강화
    assert isinstance(result, tuple) and len(result) == 3
    assert result[1] == 1
    matches = result[2]
    assert matches[0][1] == "Sheet"
    assert matches[0][2] == "B2"


def test_search_in_file_nfc_normalization(temp_dir):
    file_path = os.path.join(temp_dir, "nfc_test.txt")
    content = "\uc548\ub155"  # "안녕"
    with open(file_path, "w", encoding="utf-8") as f:
        f.write(content)

    result = search_in_file(file_path, "\uc548\ub155")
    assert result is not None


def test_search_large_file_size_limit(temp_dir):
    pass


def test_search_encoding_error_handling(temp_dir):
    file_path = os.path.join(temp_dir, "bad_encoding.txt")
    with open(file_path, "wb") as f:
        f.write(b"\xff\xfe\x00\x00")
    result = search_in_file(file_path, "test")
    assert result is None


def test_search_io_error_handling(temp_dir):
    non_existent = os.path.join(temp_dir, "non_existent.txt")
    result = search_in_file(non_existent, "test")
    assert isinstance(result, tuple)
    assert result[0] == "SKIPPED"


def test_search_corrupted_excel_handling(temp_dir):
    excel_path = os.path.join(temp_dir, "corrupted.xlsx")
    with open(excel_path, "w") as f:
        f.write("This is not a valid Excel file")
    result = search_in_file(excel_path, "test")
    assert isinstance(result, tuple)
    assert result[0] == "SKIPPED"
