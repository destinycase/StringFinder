"""
[test_match_precision.py]

이 테스트는 다양한 특수 파일 형식 및 검색 모드에서의 검색 결과 '정밀도'를 검증합니다.

- 테스트 목적:
  1. 부분 일치(Partial)와 전체 일치(Exact) 모드 간의 명확한 검색 결과 차별화 보장.
  2. JSON, XML, Excel, Archive 파일 등 구조화된 데이터 내의 키워드 추출 정확도 확인.

- 주요 검증 사항:
  1. 텍스트 파일 내 단어 경계 및 완전 일치 매칭 로직.
  2. JSON/XML 데이터 트리 순회 시 키/값 영역의 정밀한 키워드 필터링.
  3. Archive 특수 포맷의 계층적 구조 내 검색 결과 위치 정보 정확도.
"""

import json
from typing import cast
from unittest.mock import patch

import openpyxl

from core.search_engine import SearchResult, search_in_file
from sf_utils.constants import Constants


def _require_search_result(result: object) -> SearchResult:
    assert result is not None
    assert isinstance(result, tuple) and len(result) == 3
    return cast(SearchResult, result)


def test_normal_text_precision(tmp_path):
    file_path = tmp_path / "test.txt"
    file_path.write_text("user_id\nid\nidentity", encoding="utf-8")

    res_partial = _require_search_result(search_in_file(str(file_path), "id"))
    assert res_partial[1] == 3

    res_exact = _require_search_result(search_in_file(str(file_path), "id", special_mode=Constants.MODE_EXACT))
    assert res_exact[1] == 1
    assert res_exact[2][0][1].strip() == "id"


def test_json_special_precision(tmp_path):
    data = {"user": "admin", "user_id": "admin_123", "description": "This is admin user"}
    file_path = tmp_path / "test.json"
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(data, f)

    res_partial = _require_search_result(search_in_file(str(file_path), "admin", special_mode=Constants.MODE_JSON))
    assert res_partial[1] == 3

    res_exact = _require_search_result(
        search_in_file(str(file_path), "admin", special_mode=f"{Constants.MODE_JSON} {Constants.MODE_EXACT}")
    )
    assert res_exact[1] == 1
    assert res_exact[2][0][2] == "admin"


def test_xml_special_precision(tmp_path):
    xml_content = """<root>
        <item name="target">target_value</item>
        <item name="target_extra">extra</item>
        <note>target</note>
    </root>"""
    file_path = tmp_path / "test.xml"
    file_path.write_text(xml_content, encoding="utf-8")

    res_partial = _require_search_result(search_in_file(str(file_path), "target", special_mode=Constants.MODE_XML))
    assert res_partial[1] == 4

    res_exact = _require_search_result(
        search_in_file(str(file_path), "target", special_mode=f"{Constants.MODE_XML} {Constants.MODE_EXACT}")
    )
    assert res_exact[1] == 2
    for match in res_exact[2]:
        content = str(match[2]).lower()
        assert "target" in content
        assert "target_value" not in content


def test_excel_precision(tmp_path):
    file_path = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Match"
    ws["A2"] = "Matching"
    ws["B1"] = "Match_Extra"
    wb.save(file_path)

    res_partial = _require_search_result(search_in_file(str(file_path), "Match"))
    assert res_partial[1] == 3

    res_exact = _require_search_result(search_in_file(str(file_path), "Match", special_mode=Constants.MODE_EXACT))
    assert res_exact[1] == 1
    assert "A1" in str(res_exact[2][0][2])


def test_archive_precision(tmp_path):
    data = {
        "Subnamespaces": [
            {
                "Namespace": "Game",
                "Children": [
                    {"Key": "K1", "Source": {"Text": "Start"}, "Translation": {"Text": "Begin"}},
                    {"Key": "K2", "Source": {"Text": "Restart"}, "Translation": {"Text": "Again"}},
                ],
            }
        ]
    }
    file_path = tmp_path / "test.archive"
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(data, f)

    with patch("core.search_engine.HAS_RUST_ENGINE", False):
        res_partial = _require_search_result(
            search_in_file(str(file_path), "Start", special_mode=Constants.MODE_ARCHIVE)
        )
        assert res_partial[1] == 2

        res_exact = _require_search_result(
            search_in_file(str(file_path), "Start", special_mode=f"{Constants.MODE_ARCHIVE} {Constants.MODE_EXACT}")
        )
        assert res_exact[1] == 1
        assert res_exact[2][0][3] == "Start"
